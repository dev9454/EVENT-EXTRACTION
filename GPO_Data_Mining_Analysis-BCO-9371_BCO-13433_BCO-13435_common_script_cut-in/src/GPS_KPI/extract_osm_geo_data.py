# Created 21.08.2024 by Jakub Grzesiak
# Last modified: 30.09.2024 by Jakub Grzesiak
import os
import sys
import math
import time
import argparse
import openpyxl
import numpy as np
import pandas as pd
import geopandas as gpd
from openpyxl import load_workbook
from shapely.geometry import Point, LineString, MultiLineString, Polygon

sys.path.insert(1, os.path.dirname(os.path.realpath(__file__)))
from extractor_common_functions import ExtractorCommonFunctions


class ExtractOsmGeoData:
    geo_data_objects = ['road_type',
                        'barrier',
                        'bridges',
                        'buildings',
                        'bus_stop',
                        'crossing',
                        'gantry',
                        'manhole',
                        'parking',
                        'power_line',
                        'railway_crossing',
                        'roundabout',
                        'toll_booth',
                        'traffic_calming',
                        'traffic_sign',
                        'traffic_signals',
                        'tunnel']

    def __init__(self):
        self._func_name = os.path.basename(__file__).replace(".py", "")
        self._version = '1.0'
        self._venv = ExtractorCommonFunctions.venv_path
        self._red_venv = self._venv
        self._headers = dict()

        self._headers['barrier'] = ['log_path',
                                    'log_name',
                                    'region',
                                    'event_start_ctime',
                                    'event_end_ctime',
                                    'duration',
                                    'milage',
                                    'el_id',
                                    'type',
                                    'object_name',
                                    'object_timestamp',
                                    'barrier',
                                    'geometry',
                                    'geo_buffer',
                                    'matched_object']

        self._headers['bridges'] = ['log_path',
                                    'log_name',
                                    'region',
                                    'event_start_ctime',
                                    'event_end_ctime',
                                    'duration',
                                    'milage',
                                    'el_id',
                                    'type',
                                    'object_name',
                                    'object_timestamp',
                                    'bridge',
                                    'highway',
                                    'lanes',
                                    'layer',
                                    'maxspeed',
                                    'name',
                                    'name:en',
                                    'oneway',
                                    'surface',
                                    'bicycle',
                                    'foot',
                                    'covered',
                                    'hov',
                                    'hov:lanes',
                                    'bus:lanes',
                                    'access:lanes',
                                    'motorcycle:lanes',
                                    'turn:lanes',
                                    'geometry',
                                    'geo_buffer',
                                    'matched_object']  # + milage

        self._headers['buildings'] = ['log_path',
                                      'log_name',
                                      'region',
                                      'event_start_ctime',
                                      'event_end_ctime',
                                      'duration',
                                      'el_id',
                                      'type',
                                      'object_name',
                                      'object_timestamp',
                                      'amenity',
                                      'building',
                                      'name',
                                      'name:en',
                                      'building:levels',
                                      'alt_name',
                                      'school',
                                      'geometry',
                                      'geo_buffer',
                                      'matched_object']

        self._headers['bus_stop'] = ['log_path',
                                     'log_name',
                                     'region',
                                     'event_start_ctime',
                                     'event_end_ctime',
                                     'duration',
                                     'el_id',
                                     'type',
                                     'object_name',
                                     'object_timestamp',
                                     'highway',
                                     'bus_stop',
                                     'name',
                                     'name:en',
                                     'geometry',
                                     'geo_buffer',
                                     'matched_object']

        self._headers['crossing'] = ['log_path',
                                     'log_name',
                                     'region',
                                     'event_start_ctime',
                                     'event_end_ctime',
                                     'duration',
                                     'el_id',
                                     'type',
                                     'object_name',
                                     'object_timestamp',
                                     'highway',
                                     'crossing',
                                     'crossing:island',
                                     'crossing:markings',
                                     'crossing:signals',
                                     'geometry',
                                     'geo_buffer',
                                     'matched_object']

        self._headers['gantry'] = ['log_path',
                                   'log_name',
                                   'region',
                                   'event_start_ctime',
                                   'event_end_ctime',
                                   'duration',
                                   'el_id',
                                   'type',
                                   'object_name',
                                   'object_timestamp',
                                   'man_made',
                                   'gantry',
                                   'gantry:type',
                                   'traffic_signals',
                                   'destionation_sign',
                                   'highway',
                                   'geometry',
                                   'geo_buffer',
                                   'matched_object']

        self._headers['manhole'] = ['log_path',
                                    'log_name',
                                    'region',
                                    'event_start_ctime',
                                    'event_end_ctime',
                                    'duration',
                                    'el_id',
                                    'type',
                                    'object_name',
                                    'object_timestamp',
                                    'manhole',
                                    'geometry',
                                    'geo_buffer',
                                    'matched_object']

        self._headers['parking'] = ['log_path',
                                    'log_name',
                                    'region',
                                    'event_start_ctime',
                                    'event_end_ctime',
                                    'duration',
                                    'el_id',
                                    'type',
                                    'object_name',
                                    'object_timestamp',
                                    'parking',
                                    'surface',
                                    'street_side',
                                    'multi-storey',
                                    'underground',
                                    'lane',
                                    'layby',
                                    'on_kerb',
                                    'rooftop',
                                    'name',
                                    'name:en',
                                    'geometry',
                                    'geo_buffer',
                                    'matched_object']

        self._headers['power_line'] = ['log_path',
                                       'log_name',
                                       'region',
                                       'event_start_ctime',
                                       'event_end_ctime',
                                       'duration',
                                       'el_id',
                                       'type',
                                       'object_name',
                                       'object_timestamp',
                                       'cables',
                                       'circuits',
                                       'frequency',
                                       'name',
                                       'operator',
                                       'power',
                                       'voltage',
                                       'layer',
                                       'line',
                                       'geometry',
                                       'geo_buffer',
                                       'matched_object']

        self._headers['railway_crossing'] = ['log_path',
                                             'log_name',
                                             'region',
                                             'event_start_ctime',
                                             'event_end_ctime',
                                             'duration',
                                             'el_id',
                                             'type',
                                             'object_name',
                                             'object_timestamp',
                                             'railway',
                                             'geometry',
                                             'geo_buffer',
                                             'matched_object']

        self._headers['road_type'] = ['log_path',
                                      'log_name',
                                      'region',
                                      'event_start_ctime',
                                      'event_end_ctime',
                                      'duration',
                                      'milage',
                                      'el_id',
                                      'type',
                                      'object_name',
                                      'object_timestamp',
                                      'covered',
                                      'highway',
                                      'incline',
                                      'lanes',
                                      'layer',
                                      'level',
                                      'maxspeed',
                                      'name',
                                      'name:en',
                                      'oneway',
                                      'tunnel',
                                      'surface',
                                      'bicycle',
                                      'foot',
                                      'crossing',
                                      'bridge',
                                      'hov',
                                      'hov:lanes',
                                      'bus:lanes',
                                      'access:lanes',
                                      'motorcycle:lanes',
                                      'turn:lanes',
                                      'traffic_sign',
                                      'busway',
                                      'construction',
                                      'geometry',
                                      'geo_buffer',
                                      'matched_object']  # + milage

        self._headers['roundabout'] = ['log_path',
                                       'log_name',
                                       'region',
                                       'event_start_ctime',
                                       'event_end_ctime',
                                       'duration',
                                       'el_id',
                                       'type',
                                       'object_name',
                                       'object_timestamp',
                                       'highway',
                                       'junction',
                                       'geometry',
                                       'geo_buffer',
                                       'matched_object']

        self._headers['toll_both'] = ['log_path',
                                      'log_name',
                                      'region',
                                      'event_start_ctime',
                                      'event_end_ctime',
                                      'duration',
                                      'el_id',
                                      'type',
                                      'object_name',
                                      'object_timestamp',
                                      'barrier',
                                      'name',
                                      'geometry',
                                      'geo_buffer',
                                      'matched_object']

        self._headers['traffic_calming'] = ['log_path',
                                            'log_name',
                                            'region',
                                            'event_start_ctime',
                                            'event_end_ctime',
                                            'duration',
                                            'el_id',
                                            'type',
                                            'object_name',
                                            'object_timestamp',
                                            'traffic_calming',
                                            'direction',
                                            'surface',
                                            'geometry',
                                            'geo_buffer',
                                            'matched_object']

        self._headers['traffic_sign'] = ['log_path',
                                         'log_name',
                                         'region',
                                         'event_start_ctime',
                                         'event_end_ctime',
                                         'duration',
                                         'el_id',
                                         'type',
                                         'object_name',
                                         'object_timestamp',
                                         'highway',
                                         'name',
                                         'maxspeed',
                                         'geometry',
                                         'geo_buffer',
                                         'matched_object']

        self._headers['traffic_signals'] = ['log_path',
                                            'log_name',
                                            'region',
                                            'event_start_ctime',
                                            'event_end_ctime',
                                            'duration',
                                            'el_id',
                                            'type',
                                            'object_name',
                                            'object_timestamp',
                                            'highway',
                                            'traffic_signals',
                                            'traffic_signals:direction',
                                            'crossing',
                                            'geometry',
                                            'geo_buffer',
                                            'matched_object']

        self._headers['tunnel'] = ['log_path',
                                   'log_name',
                                   'region',
                                   'event_start_ctime',
                                   'event_end_ctime',
                                   'duration',
                                   'milage',
                                   'el_id',
                                   'type',
                                   'object_name',
                                   'object_timestamp',
                                   'covered',
                                   'highway',
                                   'incline',
                                   'lanes',
                                   'layer',
                                   'level',
                                   'maxspeed',
                                   'name',
                                   'name:en',
                                   'oneway',
                                   'tunnel',
                                   'surface',
                                   'bicycle',
                                   'foot',
                                   'hov',
                                   'hov:lanes',
                                   'bus:lanes',
                                   'access:lanes',
                                   'motorcycle:lanes',
                                   'turn:lanes',
                                   'traffic_sign',
                                   'geometry',
                                   'geo_buffer',
                                   'matched_object']  # + milage

        self.osm_data_dir = r"/mnt/usmidet/projects/STLA-THUNDER/8-Users/AlgoGroup/GeoData/RoadScenarios_GPKG_files"

        self.geo_data_objects = self.__class__.geo_data_objects

        self.standard_headers = ['log_path',
                                 'log_name',
                                 'region',
                                 'event_start_ctime',
                                 'event_end_ctime',
                                 'duration']

        self.regions = ['us-northeast',
                        'us-midwest',
                        'us-west',
                        'us-south',
                        'japan',
                        'gcc-states'
                        ]

        # self.object_map_color = {'bridges': 'red',
        #                          'road_type': 'blue',
        #                          'tunnel': 'green',
        #                          'barrier': 'cyan',
        #                          'power_line': 'magenta',
        #                          'railway_crossing': 'deepskyblue',
        #                          'toll_booth': 'white',
        #                          'traffic_calming': 'grey',
        #                          'crossing': 'lightcoral',
        #                          'traffic_signals': 'olive',
        #                          'roundabout': 'teal'}


    def get_version(self):
        """
        :return: Return script version
        """
        return self._version

    def get_func_name(self):
        """
        :return: Returns function name
        """
        return self._func_name

    def get_headers(self):
        """
        :return: returns headers
        """
        return self._headers

    def get_min_value(self, src_clm: pd.Series) -> float:
        """
        Function returns min value.
        """
        return src_clm.min()

    def get_std(self, src_clm: pd.Series) -> float:
        """
        Function returns std value.
        """
        return src_clm.std()

    def get_mean(self, src_clm: pd.Series) -> float:
        """
        Function returns mean value.
        """
        return src_clm.mean()

    def get_venv(self):
        return self._venv

    def get_red_venv(self):
        return self._red_venv

    def generate_plot(self):
        """
        Uncomment this method if output dir value is required for generating plot.
        kwargs["plot_dir"] will contain output dir value. kwargs available in run method below.
        :return: Boolean True
        """
        return True

    @staticmethod
    def get_bbox(gps_df):
        buffer = 0.0000449660803  # ~5m
        # buffer = 0.0000224830402  # ~2.5m
        bbox = gpd.GeoDataFrame(gps_df, geometry='geometry').total_bounds.tolist()
        bbox = [coord - buffer if idx < 2 else coord + buffer for idx, coord in enumerate(bbox)]
        return bbox

    def get_osm_data(self, osm_file_name, bbox=None):
        import glob
        try:
            file_path = glob.glob(os.path.join(self.osm_data_dir, osm_file_name))[0]
            if bbox:
                return gpd.read_file(file_path, bbox=bbox)
            else:
                return gpd.read_file(file_path)
        except Exception as ex:
            print(f'Exception raised in "get_gsm_data" (line 504): {ex}')
            return gpd.GeoDataFrame()

    @staticmethod
    def get_gdf(gps_df):
        # Create route as lines from point to point (last row excluded)
        gps_line_route = [LineString([Point(gps_df.long_pos[row], gps_df.lat_pos[row]),
                                      Point(gps_df.long_pos[row + 1], gps_df.lat_pos[row + 1])]) for row in
                          range(len(gps_df) - 1)]

        # drop last row
        gps_df.drop(gps_df.tail(1).index, inplace=True)

        # update gps dataframw with the route
        gps_df['geometry'] = gps_line_route

        # create geodataframe
        return gpd.GeoDataFrame(gps_df, geometry='geometry', crs=4326)

    def get_gdf_event(self, lat_gps, long_gps, angle, object_buffer=3.5, polygon_size=100):
        df = pd.DataFrame(columns=['lat_pos', 'long_pos', 'gps_heading'],
                          data=[[lat_gps, long_gps, angle]])

        df['geometry'] = Point(long_gps, lat_gps)
        gdf = gpd.GeoDataFrame(df, geometry='geometry', crs=4326)
        gdf['geo_buffer'] = gdf.to_crs(3035).buffer(object_buffer, cap_style='square').to_crs(4326)

        # bbox
        bbox = gdf['geo_buffer'].bounds.values.tolist()[0]
        polygon_bbox = gdf.to_crs(3035).buffer(polygon_size, cap_style='square').to_crs(4326).bounds.values.tolist()[0]

        # empty objects
        objects = {geo_object: False for geo_object in self.geo_data_objects}

        return gdf, bbox, polygon_bbox, objects

    @staticmethod
    def get_target_coordinates(host_lat_gps, host_long_gps, heading_angle, target_lat_pos, target_long_pos):
        import geopy
        from geopy.distance import distance
        # Host and Target position as Point object
        host_position = geopy.Point(host_lat_gps, host_long_gps)
        # Calculate the true bearing
        real_heading = (heading_angle + 90) % 360
        bearing = (real_heading + math.degrees(math.atan2(target_lat_pos, target_long_pos))) % 360

        # Calculate the distance between host and target
        distance_to_target = math.sqrt(target_long_pos ** 2 + target_lat_pos ** 2)

        # Calculate the new target position using the bearing and distance
        target_position = distance(meters=distance_to_target).destination(host_position, bearing)

        return target_position.latitude, target_position.longitude, bearing

    @staticmethod
    def line_to_polygon(geometry):
        if geometry.is_ring:
            return Polygon(geometry)
        else:
            return geometry

    def check_geo_data(self, host_lat_gps, host_long_gps, heading_angle, target_lat_pos, target_long_pos):
        object_buffer = 3.5
        try:
            host_gdf, host_bbox, host_polygon, host_objects = self.get_gdf_event(host_lat_gps, host_long_gps,
                                                                                 heading_angle, object_buffer)
            target_lat_gps, target_long_gps, bearing = self.get_target_coordinates(host_lat_gps, host_long_gps,
                                                                                   heading_angle, target_lat_pos,
                                                                                   target_long_pos)
            target_gdf, target_bbox, target_polygon, target_objects = self.get_gdf_event(target_lat_gps, target_long_gps,
                                                                                         bearing, object_buffer)

            regions = self.regions
            for object_type in self.geo_data_objects:

                for region in regions:
                    osm_file_name = region + "-*_" + object_type + ".gpkg"
                    osm_data_host = self.get_osm_data(osm_file_name, bbox=host_polygon)
                    osm_data_host['geometry'] = osm_data_host.geometry.apply(
                        self.line_to_polygon)  # Convert closed line to polygon
                    osm_data_target = self.get_osm_data(osm_file_name, bbox=target_polygon)
                    osm_data_target['geometry'] = osm_data_target.geometry.apply(
                        self.line_to_polygon) # Convert closed line to polygon

                    for _, row in osm_data_host.iterrows():
                        if any(host_gdf['geo_buffer'].intersects(row['geometry'])):
                            if object_type == 'road_type':
                                host_objects[object_type] = row['highway']
                            else:
                                host_objects[object_type] = True
                            break

                    for _, row in osm_data_target.iterrows():
                        if any(target_gdf['geo_buffer'].intersects(row['geometry'])):
                            if object_type == 'road_type':
                                target_objects[object_type] = row['highway']
                            else:
                                target_objects[object_type] = True
                            break

                    if not osm_data_host.empty or not osm_data_target.empty:
                        regions = [region]
                        break

            return list(host_objects.values()), list(target_objects.values())

        except Exception as ex:
            print(f'Exception raised in "check_geo_data" (line 565): {ex}')
            return ([False for _ in self.geo_data_objects],
                    [False for _ in self.geo_data_objects])

    @staticmethod
    def counter_bar(cnt: int, t0: float, file: pd.DataFrame) -> int:
        tc = time.time()
        delta_t = (tc - t0) / cnt
        estymat_t = time.strftime('%H:%M:%S', time.gmtime((file.shape[0] - cnt) * delta_t))

        print(f'Progress: [{"=" * int(cnt / file.shape[0] * 49)}>{" " * (49 - int(cnt / file.shape[0] * 49))}]' +
              f'{" " * 3}[{cnt}/{file.shape[0]} {round(cnt / file.shape[0] * 100, 2)}%]{f" " * 3}[~ {estymat_t}]',
              end='\r')
        return cnt + 1

    def check_geo_data_in_extractor(self, file_name, sheet_name):
        extractor = pd.read_excel(file_name, sheet_name=sheet_name)
        workbook = load_workbook(file_name)
        writer = pd.ExcelWriter(file_name.replace('.xlsx', '_new.xlsx'), engine='openpyxl')
        writer.sheets.update({x.title: x for x in workbook.worksheets})

        host_columns = [f"host_{geo_object}" for geo_object in self.geo_data_objects]
        target_columns = [f"target_{geo_object}" for geo_object in self.geo_data_objects]
        extractor[host_columns] = ''
        extractor[target_columns] = ''
        gps_lat = extractor['gps_lat']
        gps_long = extractor['gps_long']
        if 'rt1_drop' in file_name:
            gps_heading = extractor['gps_heading']
            target_lat = extractor['lat_pos']
            target_long = extractor['long_pos']
        elif 'rt1_radar_only' in file_name:
            gps_heading = extractor['gps_heading']
            target_lat = extractor['rt1_lat_pos']
            target_long = extractor['rt1_long_pos']
        elif 'AEB_events' in file_name:
            gps_heading = extractor['gps_heading_deg']
            target_lat = extractor['lat_pos']
            target_long = extractor['long_pos']
        else:
            print('Extractor not supported')
            sys.exit()

        cnt = 1
        t0 = time.time()
        for index, row in extractor.iterrows():
            cnt = self.counter_bar(cnt=cnt, t0=t0, file=extractor)
            host_objects, target_objects = self.check_geo_data(gps_lat[index],
                                                               gps_long[index],
                                                               gps_heading[index],
                                                               target_lat[index],
                                                               target_long[index])

            extractor.loc[index, host_columns] = host_objects
            extractor.loc[index, target_columns] = target_objects

        extractor.to_excel(writer, sheet_name=sheet_name, index=False)

        writer.save()
        writer.close()

    def kpi_sheet_generation(self, output_excel_sheet):
        """
        This function will be called from write_excel_to_excel and take out excel as input and add KPI sheet to it.
        :param output_excel_sheet: excel sheet generated after reducer.
        :return:
        """
        workbook_df = pd.read_excel(output_excel_sheet, sheet_name=None)
        workbook = load_workbook(output_excel_sheet)
        writer = pd.ExcelWriter(output_excel_sheet, engine='openpyxl', mode='a')

        writer.sheets.update({x.title: x for x in workbook.worksheets})

        kpi_overall = pd.DataFrame(index=self.geo_data_objects, columns=['count', 'milage', 'duration'])

        for geo_object, df in workbook_df.items():
            if geo_object == 'About':
                continue
            try:
                df = df.fillna('N/A')

                if geo_object == 'barrier':
                    count = df.groupby(['barrier']).size().rename('count', inplace=True)
                    milage = df.groupby(['barrier'])['milage'].sum()
                    duration = df.groupby(['barrier'])['duration'].sum().rename('hours',
                                                                                inplace=True) / 3600  # seconds -> hours
                    kpi_df = pd.concat([count, milage, duration], axis=1)
                elif geo_object == 'bridges' or geo_object == 'road_type' or geo_object == 'tunnel':
                    count = df.groupby(['highway', 'surface', 'lanes']).size().rename('count', inplace=True)
                    milage = df.groupby(['highway', 'surface', 'lanes'])['milage'].sum()
                    duration = df.groupby(['highway', 'surface', 'lanes'])['duration'].sum().rename('hours',
                                                                                                    inplace=True) / 3600  # seconds -> hours
                    kpi_df = pd.concat([count, milage, duration], axis=1)
                elif geo_object == 'buildings':
                    kpi_df = df.groupby(['building']).size().rename('count', inplace=True)
                elif (geo_object == 'bus_stop' or geo_object == 'gantry' or geo_object == 'manhole'
                      or geo_object == 'parking' or geo_object == 'traffic_calming'):
                    kpi_df = df.groupby([geo_object]).size().rename('count', inplace=True)
                elif geo_object == 'crossing':
                    kpi_df = df.groupby(
                        ['crossing', 'crossing:island', 'crossing:markings', 'crossing:signals']).size().rename('count',
                                                                                                                inplace=True)
                elif geo_object == 'power_line':
                    kpi_df = df.groupby(['power', 'voltage', 'cables', 'frequency']).size().rename('count',
                                                                                                   inplace=True)
                elif geo_object == 'railway_crossing':
                    kpi_df = df.groupby(['railway']).size().rename('count', inplace=True)
                elif geo_object == 'roundabout' or geo_object == 'traffic_sign':
                    kpi_df = df.groupby(['highway']).size().rename('count', inplace=True)
                elif geo_object == 'toll_booth':
                    kpi_df = df.groupby(['barrier']).size().rename('count', inplace=True)
                elif geo_object == 'traffic_signals':
                    kpi_df = df.groupby(['traffic_signals', 'traffic_signals:direction', 'crossing']).size().rename(
                        'count', inplace=True)
                else:
                    continue

                kpi_df.to_excel(writer, index_label=kpi_df.index.names, sheet_name=f'{geo_object}_KPI')

                if geo_object == 'bridges' or geo_object == 'road_type' or geo_object == 'tunnel' or geo_object == 'barrier':
                    kpi_overall.loc[geo_object, :] = kpi_df.sum()
                else:
                    kpi_overall.loc[geo_object, 'count'] = kpi_df.values.sum()

            except Exception as ex:
                print(ex)
                kpi_overall.loc[geo_object, 'count'] = 0

        kpi_overall.to_excel(writer, index_label=kpi_overall.index.names, sheet_name='KPI', na_rep='N/A')

        writer.close()

    def run(self, file_name, **kwargs):
        import folium
        try:

            if 'map_plot_dir' in kwargs.keys() and kwargs['map_plot_dir']:
                if not os.path.isdir(kwargs['map_plot_dir']):
                    os.mkdir(kwargs['map_plot_dir'])
                plot_dir = kwargs['rt_target_plot_dir']
            elif 'plot_dir' in kwargs.keys() and kwargs['plot_dir']:
                plot_dir = kwargs['plot_dir']
            else:
                plot_dir = os.path.dirname(file_name)

            object_buffer = 0  # dla różnych obiektów ustalić średnie wartości, most może być za duży; multiply by 'lanes' value
            osm_buffer = 3.5  # 3,5m

            log_path, log_name = os.path.split(file_name)
            extractor = ExtractorCommonFunctions(file_name)
            data = extractor.read_mat()
            gps_df = extractor.extract_gps_position_df()

            gdf = self.get_gdf(gps_df)

            bbox_area = self.get_bbox(gps_df)  # Check if buffer is also needed

            # create buffer for route
            gdf['geo_buffer'] = gdf.to_crs(3035).buffer(object_buffer, cap_style='flat', join_style='mitre').to_crs(
                4326)

            V = abs(data['mudp']['Feature_20ms']['Host_Vehicle_State_Stream']['Host_Speed_mps'])
            a = abs(data['mudp']['Feature_20ms']['Host_Vehicle_State_Stream']['Host_Long_Accel_mpss'])
            t = data['mudp']['Feature_20ms']['header']['time'][:, 0]
            delta_t = abs(np.diff(t))

            out = dict()

            output_map = gdf.explore(name='log route')

            for object_type in self.geo_data_objects:
                complete_events = list()

                for region in self.regions:
                    osm_file_name = region + "-*_" + object_type + ".gpkg"
                    osm_data = self.get_osm_data(osm_file_name, bbox=bbox_area)
                    osm_data['color'] = 'red'

                    if osm_data.empty:
                        continue
                    else:
                        osm_data['geometry'] = osm_data.geometry.apply(self.line_to_polygon)  # Convert closed line to polygon
                        osm_data['object_timestamp'] = osm_data['object_timestamp'].astype(str)
                    osm_data['geo_buffer'] = osm_data.to_crs(3035).buffer(osm_buffer, cap_style='round').to_crs(
                        4326)  # 'flat', join_style='mitre'

                    for index, row in osm_data.iterrows():
                        if object_buffer == 0:
                            event_idx = list(np.where(gdf['geometry'].intersects(row['geo_buffer']))[0])
                        else:
                            event_idx = list(np.where(gdf['geo_buffer'].intersects(row['geo_buffer']))[0])
                        if not event_idx:
                            continue
                        else:
                            osm_data.loc[index, 'color'] = 'green'

                        # list of falling and rising edges [[start index, end index]]
                        transition_idx = list(zip([x for x in event_idx if x - 1 not in event_idx],
                                                  [x for x in event_idx if x + 1 not in event_idx]))

                        for start_idx, end_idx in transition_idx:
                            if object_type in ['bridges', 'tunnel', 'road_type', 'barrier']:

                                s = (np.abs(t - gdf['time'][start_idx])).argmin()
                                e = (np.abs(t - gdf['time'][end_idx])).argmin()
                                # (V * t) + (a/2 * t**2)
                                milage = sum((V[s:e] * delta_t[s:e]) + (a[s:e] / 2 * (delta_t[s:e]**2))) * 0.000621371
                                # c = gdf['milage'][start_idx:end_idx].sum()
                                complete_events += [[
                                                        log_path,
                                                        log_name,
                                                        region,
                                                        gdf['time'][start_idx],
                                                        gdf['time'][end_idx],
                                                        gdf['time'][end_idx] - gdf['time'][start_idx],
                                                        milage] + \
                                                    row.drop('color').to_list() + \
                                                    [MultiLineString(list(gdf['geometry'][start_idx:end_idx + 1]))]]
                            else:
                                complete_events += [[
                                                        log_path,
                                                        log_name,
                                                        region,
                                                        gdf['time'][start_idx],
                                                        gdf['time'][end_idx],
                                                        gdf['time'][end_idx] - gdf['time'][start_idx]] + \
                                                    row.drop('color').to_list() + \
                                                    [MultiLineString(list(gdf['geometry'][start_idx:end_idx + 1]))]]

                    osm_data['geometry'] = osm_data['geo_buffer']
                    output_map = osm_data.explore(m=output_map,
                                                  name=object_type,
                                                  color=osm_data['color'])

                    continue

                if complete_events:
                    out[object_type] = np.array(complete_events)

            folium.LayerControl(collapsed=False).add_to(output_map)
            plot_file = os.path.join(plot_dir, os.path.basename(file_name)).replace('.mat', '_osm_map_plot.html')
            output_map.save(plot_file)

            return out

        except Exception:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            return exc_obj.args[0] + " FOUND IN LINE: " + str(exc_tb.tb_lineno)


if __name__ == '__main__':
    # file_name = r"C:\Users\bjrlhg\Desktop\Darek\TNDR1_DRUK_20240722_191920_WDC5_dma_0008.mat"
    # file_name = r"C:\Users\bjrlhg\Desktop\Darek\TNDR1_KALU_20240702_134431_WDC5_dma_0006.mat"
    # file_name = r"C:\Users\bjrlhg\Desktop\Darek\TNDR1_KALU_20240702_172133_WDC5_dma_0007.mat"
    # file_name = r"C:\Users\bjrlhg\Desktop\Darek\TNDR1_KALU_20240703_141819_WDC5_dma_0042.mat"
    # kwargs = dict()
    #
    # obj = ExtractGeoData()
    # obj.osm_data_dir = r"C:\Users\bjrlhg\Desktop\GPKG_files"
    # cell = obj.run(file_name, **kwargs)
    # print(cell)
    # output_path = r"C:\Users\bjrlhg\Desktop\geo_data.xlsx"
    # with pd.ExcelWriter(output_path) as writer:
    #     for sheet in obj.get_headers():
    #         try:
    #             dfa = pd.DataFrame(cell[sheet], columns=obj.get_headers()[sheet])
    #             dfa.columns = obj._headers[sheet]
    #             dfa.to_excel(writer, sheet_name=sheet, index=False)
    #         except Exception as error:
    #             exc_type, exc_obj, exc_tb = sys.exc_info()
    #             error_msg = "ERROR in " + sheet + " tab saving" + str(error) + " FOUND IN LINE: " + str(
    #                 exc_tb.tb_lineno)
    #             print(error, error_msg)
    # print('Output Excel Saved', output_path)

    ############
    # Run script with already existing event extractor (geo_data)
    ############
    # A = ExtractGeoData()
    # A.osm_data_dir = r"C:\Users\bjrlhg\Desktop\GPKG_files"
    # A.check_geo_data( 42.36617793381478, -71.05851842242934,90, -3, 50)
    # A.check_geo_data(42.6063075155, -83.15541117766666, 358.277, 0.0, 100)
    # A.check_geo_data(42.361709, -71.077480, 5, 0, 100)

    ############
    # Run script with already existing event extractor (RT1 drop / RT1 radar only)
    ############
    ap = argparse.ArgumentParser(description='add GPS object info')
    ap.add_argument("extractor_name", type=str, help="Path event extractor file")
    ap.add_argument("-s", "--sheet_name", type=str, help="Name of worksheet", default=None)
    kwargs = vars(ap.parse_args())
    #
    extractor_name = kwargs["extractor_name"]
    sheet_name = kwargs['sheet_name'] if kwargs['sheet_name'] else 'output'
    # extractor_name = r"C:\Users\bjrlhg\Desktop\extract_rt1_radar_only_thunder_2024-09-26_03-31-07_0.xlsx"
    # sheet_name = 'Sheet1'
    #
    A = ExtractOsmGeoData()
    # A.osm_data_dir = r"C:\Users\bjrlhg\Desktop\GPKG_files"
    A.check_geo_data_in_extractor(extractor_name, sheet_name)

    ###########
    # A = ExtractGeoData()
    # extractor_name = r"C:\Users\bjrlhg\Desktop\extract_geo_data_2024-09-27_05-37-51_0.xlsx"
    # A.kpi_sheet_generation(extractor_name)

